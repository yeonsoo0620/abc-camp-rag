"""Yes24 IT 모바일 베스트셀러 Excel 대시보드 생성 스크립트."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "yes24_it_mobile_bestsellers.csv")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "yes24_dashboard.xlsx")

# ── Styles ──
HF = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
HFL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TF = Font(name="맑은 고딕", bold=True, size=16, color="1F3864")
SF = Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
KV = Font(name="맑은 고딕", bold=True, size=18, color="1F3864")
KL = Font(name="맑은 고딕", size=10, color="595959")
TB = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
S1 = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
S2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CC = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47", "264478", "9B57A0", "636363", "EB7E30"]


def sh(ws, row, mc):
    for c in range(1, mc + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HF
        cell.fill = HFL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = TB


def sd(ws, sr, er, mc):
    for r in range(sr, er + 1):
        f = S1 if (r - sr) % 2 == 0 else S2
        for c in range(1, mc + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = f
            cell.border = TB
            cell.alignment = Alignment(vertical="center")


def kpi(ws, row, col, label, value, bg):
    cl = ws.cell(row=row, column=col, value=label)
    cl.font = KL
    cl.alignment = Alignment(horizontal="center")
    cv = ws.cell(row=row + 1, column=col, value=value)
    cv.font = KV
    cv.alignment = Alignment(horizontal="center")
    fl = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    for r in [row, row + 1]:
        ws.cell(row=r, column=col).fill = fl
        ws.cell(row=r, column=col).border = TB


# ── Data ──
df = pd.read_csv(DATA_PATH, encoding="utf-8-sig")
for col in ["판매가", "정가", "판매지수"]:
    df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "", regex=False).str.strip(), errors="coerce")
df["할인율_num"] = pd.to_numeric(df["할인율"].str.replace("%", "", regex=False).str.strip(), errors="coerce")
df["출판일_연도"] = df["출판일"].str.extract(r"(\d{4})", expand=False).astype(float)

wb = Workbook()

# ═══════ Sheet 1: 대시보드 ═══════
ws = wb.active
ws.title = "대시보드"
ws.sheet_properties.tabColor = "2F5496"

ws.merge_cells("A1:H1")
ws.cell(row=1, column=1, value="Yes24 IT 모바일 베스트셀러 대시보드").font = TF
ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:H2")
ws.cell(row=2, column=1, value=f"총 {len(df):,}권 | 기준 데이터: yes24_it_mobile_bestsellers.csv").font = Font(name="맑은 고딕", size=10, color="808080")
ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

kpis = [
    ("총 도서 수", f"{len(df):,}권", "D6E4F0"),
    ("평균 판매가", f"{df['판매가'].mean():,.0f}원", "E2EFDA"),
    ("평균 판매지수", f"{df['판매지수'].mean():,.0f}", "FCE4D6"),
    ("출판사 수", f"{df['출판사'].nunique()}개", "D9E2F3"),
    ("평균 할인율", f"{df['할인율_num'].mean():.1f}%", "FFF2CC"),
    ("최고 판매지수", f"{df['판매지수'].max():,.0f}", "E2D9F3"),
]
for i, (l, v, c) in enumerate(kpis):
    kpi(ws, 4, 1 + i * 2, l, v, c)
    ws.merge_cells(start_row=4, start_column=1 + i * 2, end_row=4, end_column=2 + i * 2)
    ws.merge_cells(start_row=5, start_column=1 + i * 2, end_row=5, end_column=2 + i * 2)
ws.row_dimensions[4].height = 20
ws.row_dimensions[5].height = 35

# Top 10 판매지수
ws.cell(row=7, column=1, value="Top 10 판매지수 도서").font = SF
top10 = df.nlargest(10, "판매지수")[["도서명", "판매지수"]].iloc[::-1]
ws.cell(row=8, column=1, value="도서명").font = HF
ws.cell(row=8, column=1).fill = HFL
ws.cell(row=8, column=2, value="판매지수").font = HF
ws.cell(row=8, column=2).fill = HFL
for i, (_, r) in enumerate(top10.iterrows()):
    t = str(r["도서명"])[:30]
    ws.cell(row=9 + i, column=1, value=t)
    ws.cell(row=9 + i, column=2, value=r["판매지수"])
sd(ws, 9, 18, 2)

ch1 = BarChart()
ch1.type = "bar"
ch1.style = 10
ch1.title = "Top 10 판매지수"
ch1.width = 28
ch1.height = 14
ch1.add_data(Reference(ws, min_col=2, min_row=8, max_row=18), titles_from_data=True)
ch1.set_categories(Reference(ws, min_col=1, min_row=9, max_row=18))
ch1.series[0].graphicalProperties.solidFill = "4472C4"
ws.add_chart(ch1, "D7")

# 출판사별 도서 수
ws.cell(row=20, column=1, value="출판사별 도서 수 (상위 10)").font = SF
pub_c = df["출판사"].value_counts().nlargest(10)
ws.cell(row=21, column=1, value="출판사").font = HF
ws.cell(row=21, column=1).fill = HFL
ws.cell(row=21, column=2, value="도서 수").font = HF
ws.cell(row=21, column=2).fill = HFL
for i, (p, c) in enumerate(pub_c.items()):
    ws.cell(row=22 + i, column=1, value=p)
    ws.cell(row=22 + i, column=2, value=c)
sd(ws, 22, 31, 2)

ch2 = BarChart()
ch2.type = "col"
ch2.style = 10
ch2.title = "출판사별 도서 수 (상위 10)"
ch2.y_axis.title = "도서 수"
ch2.width = 22
ch2.height = 12
ch2.add_data(Reference(ws, min_col=2, min_row=21, max_row=31), titles_from_data=True)
ch2.set_categories(Reference(ws, min_col=1, min_row=22, max_row=31))
ch2.series[0].graphicalProperties.solidFill = "ED7D31"
ws.add_chart(ch2, "A33")

# 연도별 출판
ws.cell(row=20, column=5, value="연도별 출판 도서 수").font = SF
yc = df["출판일_연도"].dropna().astype(int).value_counts().sort_index()
ws.cell(row=21, column=5, value="연도").font = HF
ws.cell(row=21, column=5).fill = HFL
ws.cell(row=21, column=6, value="도서 수").font = HF
ws.cell(row=21, column=6).fill = HFL
for i, (y, c) in enumerate(yc.items()):
    ws.cell(row=22 + i, column=5, value=str(int(y)))
    ws.cell(row=22 + i, column=6, value=c)
sd(ws, 22, 21 + len(yc), 6)

ch3 = BarChart()
ch3.type = "col"
ch3.style = 10
ch3.title = "연도별 출판 도서 수"
ch3.y_axis.title = "도서 수"
ch3.width = 18
ch3.height = 12
ch3.add_data(Reference(ws, min_col=6, min_row=21, max_row=21 + len(yc)), titles_from_data=True)
ch3.set_categories(Reference(ws, min_col=5, min_row=22, max_row=21 + len(yc)))
ch3.series[0].graphicalProperties.solidFill = "70AD47"
ws.add_chart(ch3, "H20")

print("[1/5] 대시보드 시트 완료")

# ═══════ Sheet 2: 가격 분석 ═══════
ws2 = wb.create_sheet("가격 분석")
ws2.sheet_properties.tabColor = "70AD47"
ws2.merge_cells("A1:H1")
ws2.cell(row=1, column=1, value="가격 분석").font = TF
ws2.row_dimensions[1].height = 35

bins = [0, 10000, 15000, 20000, 25000, 30000, 50000, 100000]
lbls = ["1만원 미만", "1~1.5만", "1.5~2만", "2~2.5만", "2.5~3만", "3~5만", "5만원 이상"]
df["가격대"] = pd.cut(df["판매가"], bins=bins, labels=lbls, right=False)
pd_ = df["가격대"].value_counts().reindex(lbls)

ws2.cell(row=3, column=1, value="가격대별 도서 수").font = SF
ws2.cell(row=4, column=1, value="가격대").font = HF; ws2.cell(row=4, column=1).fill = HFL
ws2.cell(row=4, column=2, value="도서 수").font = HF; ws2.cell(row=4, column=2).fill = HFL
ws2.cell(row=4, column=3, value="비율(%)").font = HF; ws2.cell(row=4, column=3).fill = HFL
for i, l in enumerate(lbls):
    ws2.cell(row=5 + i, column=1, value=l)
    ws2.cell(row=5 + i, column=2, value=int(pd_[l]))
    ws2.cell(row=5 + i, column=3, value=round(pd_[l] / len(df) * 100, 1))
sd(ws2, 5, 11, 3)

cp = BarChart()
cp.type = "col"; cp.style = 10; cp.title = "가격대별 도서 수 분포"
cp.y_axis.title = "도서 수"; cp.width = 22; cp.height = 12
cp.add_data(Reference(ws2, min_col=2, min_row=4, max_row=11), titles_from_data=True)
cp.set_categories(Reference(ws2, min_col=1, min_row=5, max_row=11))
cp.series[0].graphicalProperties.solidFill = "5B9BD5"
ws2.add_chart(cp, "E3")

# 정가 vs 판매가
ss = 13
ws2.cell(row=ss, column=1, value="정가 vs 판매가 비교 (상위 30권)").font = SF
ws2.cell(row=ss+1, column=1, value="도서명").font = HF; ws2.cell(row=ss+1, column=1).fill = HFL
ws2.cell(row=ss+1, column=2, value="정가").font = HF; ws2.cell(row=ss+1, column=2).fill = HFL
ws2.cell(row=ss+1, column=3, value="판매가").font = HF; ws2.cell(row=ss+1, column=3).fill = HFL
t30 = df.dropna(subset=["정가", "판매가"]).nlargest(30, "판매지수")
for i, (_, r) in enumerate(t30.iterrows()):
    ws2.cell(row=ss+2+i, column=1, value=str(r["도서명"])[:35])
    ws2.cell(row=ss+2+i, column=2, value=r["정가"])
    ws2.cell(row=ss+2+i, column=3, value=r["판매가"])
sd(ws2, ss+2, ss+1+len(t30), 3)

cs = BarChart()
cs.type = "col"; cs.grouping = "stacked"; cs.style = 10
cs.title = "정가 vs 판매가 (상위 30권)"; cs.y_axis.title = "가격 (원)"
cs.width = 28; cs.height = 14
cs.add_data(Reference(ws2, min_col=2, min_row=ss+1, max_row=ss+1+len(t30)), titles_from_data=True)
cs.add_data(Reference(ws2, min_col=3, min_row=ss+1, max_row=ss+1+len(t30)), titles_from_data=True)
cs.set_categories(Reference(ws2, min_col=1, min_row=ss+2, max_row=ss+1+len(t30)))
cs.series[0].graphicalProperties.solidFill = "4472C4"
cs.series[1].graphicalProperties.solidFill = "ED7D31"
ws2.add_chart(cs, "E15")

print("[2/5] 가격 분석 시트 완료")

# ═══════ Sheet 3: 출판사 분석 ═══════
ws3 = wb.create_sheet("출판사 분석")
ws3.sheet_properties.tabColor = "ED7D31"
ws3.merge_cells("A1:H1")
ws3.cell(row=1, column=1, value="출판사 분석").font = TF
ws3.row_dimensions[1].height = 35

ps = df.groupby("출판사").agg(
    도서수=("도서명", "count"), 평균판매가=("판매가", "mean"),
    평균판매지수=("판매지수", "mean"), 총판매지수=("판매지수", "sum"),
    평균할인율=("할인율_num", "mean"),
).sort_values("총판매지수", ascending=False).head(20)

ws3.cell(row=3, column=1, value="출판사별 종합 통계 (상위 20)").font = SF
hdrs = ["출판사", "도서 수", "평균 판매가", "평균 판매지수", "총 판매지수", "평균 할인율"]
for j, h in enumerate(hdrs, 1):
    ws3.cell(row=4, column=j, value=h)
sh(ws3, 4, 6)

for i, (pn, rd) in enumerate(ps.iterrows()):
    ws3.cell(row=5+i, column=1, value=pn)
    ws3.cell(row=5+i, column=2, value=int(rd["도서수"]))
    ws3.cell(row=5+i, column=3, value=round(rd["평균판매가"]))
    ws3.cell(row=5+i, column=3).number_format = '#,##0'
    ws3.cell(row=5+i, column=4, value=round(rd["평균판매지수"]))
    ws3.cell(row=5+i, column=4).number_format = '#,##0'
    ws3.cell(row=5+i, column=5, value=round(rd["총판매지수"]))
    ws3.cell(row=5+i, column=5).number_format = '#,##0'
    ws3.cell(row=5+i, column=6, value=round(rd["평균할인율"], 1))
sd(ws3, 5, 24, 6)

c1 = BarChart()
c1.type = "bar"; c1.style = 10; c1.title = "출판사별 총 판매지수 (상위 20)"
c1.width = 28; c1.height = 16
c1.add_data(Reference(ws3, min_col=5, min_row=4, max_row=24), titles_from_data=True)
c1.set_categories(Reference(ws3, min_col=1, min_row=5, max_row=24))
c1.series[0].graphicalProperties.solidFill = "4472C4"
ws3.add_chart(c1, "A26")

c2 = BarChart()
c2.type = "col"; c2.style = 10; c2.title = "출판사별 도서 수 (상위 20)"
c2.y_axis.title = "도서 수"; c2.width = 28; c2.height = 14
c2.add_data(Reference(ws3, min_col=2, min_row=4, max_row=24), titles_from_data=True)
c2.set_categories(Reference(ws3, min_col=1, min_row=5, max_row=24))
c2.series[0].graphicalProperties.solidFill = "70AD47"
ws3.add_chart(c2, "H26")

print("[3/5] 출판사 분석 시트 완료")

# ═══════ Sheet 4: 할인율 분석 ═══════
ws4 = wb.create_sheet("할인율 분석")
ws4.sheet_properties.tabColor = "FFC000"
ws4.merge_cells("A1:H1")
ws4.cell(row=1, column=1, value="할인율 분석").font = TF
ws4.row_dimensions[1].height = 35

dc = df["할인율"].dropna().value_counts().sort_index()
ws4.cell(row=3, column=1, value="할인율 분포").font = SF
ws4.cell(row=4, column=1, value="할인율").font = HF; ws4.cell(row=4, column=1).fill = HFL
ws4.cell(row=4, column=2, value="도서 수").font = HF; ws4.cell(row=4, column=2).fill = HFL
ws4.cell(row=4, column=3, value="비율(%)").font = HF; ws4.cell(row=4, column=3).fill = HFL
for i, (d, c) in enumerate(dc.items()):
    ws4.cell(row=5+i, column=1, value=d)
    ws4.cell(row=5+i, column=2, value=c)
    ws4.cell(row=5+i, column=3, value=round(c/len(df)*100, 1))
sd(ws4, 5, 4+len(dc), 3)

dp = PieChart()
dp.title = "할인율 분포"; dp.style = 10; dp.width = 18; dp.height = 14
dp.add_data(Reference(ws4, min_col=2, min_row=4, max_row=4+len(dc)), titles_from_data=True)
dp.set_categories(Reference(ws4, min_col=1, min_row=5, max_row=4+len(dc)))
dp.dataLabels = DataLabelList()
dp.dataLabels.showPercent = True
dp.dataLabels.showVal = True
for idx in range(len(dc)):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = CC[idx % len(CC)]
    dp.series[0].data_points.append(pt)
ws4.add_chart(dp, "E3")

# 출판사별 평균 할인율
pdd = df.groupby("출판사")["할인율_num"].mean().nlargest(15).sort_values()
ds = 5 + len(dc) + 1
ws4.cell(row=ds, column=1, value="출판사별 평균 할인율 (상위 15)").font = SF
ws4.cell(row=ds+1, column=1, value="출판사").font = HF; ws4.cell(row=ds+1, column=1).fill = HFL
ws4.cell(row=ds+1, column=2, value="평균 할인율(%)").font = HF; ws4.cell(row=ds+1, column=2).fill = HFL
for i, (p, r) in enumerate(pdd.items()):
    ws4.cell(row=ds+2+i, column=1, value=p)
    ws4.cell(row=ds+2+i, column=2, value=round(r, 1))
sd(ws4, ds+2, ds+1+len(pdd), 2)

cd = BarChart()
cd.type = "bar"; cd.style = 10; cd.title = "출판사별 평균 할인율 (상위 15)"
cd.width = 22; cd.height = 12
cd.add_data(Reference(ws4, min_col=2, min_row=ds+1, max_row=ds+1+len(pdd)), titles_from_data=True)
cd.set_categories(Reference(ws4, min_col=1, min_row=ds+2, max_row=ds+1+len(pdd)))
cd.series[0].graphicalProperties.solidFill = "FFC000"
ws4.add_chart(cd, "A" + str(ds + 2 + len(pdd) + 1))

print("[4/5] 할인율 분석 시트 완료")

# ═══════ Sheet 5: 원본 데이터 ═══════
ws5 = wb.create_sheet("원본 데이터")
ws5.sheet_properties.tabColor = "808080"
orig_cols = ["순위", "도서번호", "도서명", "링크", "저자", "출판사", "출판일", "판매가", "정가", "할인율", "판매지수"]
for j, h in enumerate(orig_cols, 1):
    ws5.cell(row=1, column=j, value=h)
sh(ws5, 1, len(orig_cols))
for i, (_, r) in enumerate(df.iterrows()):
    for j, c in enumerate(orig_cols, 1):
        ws5.cell(row=2+i, column=j, value=r[c])
sd(ws5, 2, 1+len(df), len(orig_cols))
ws5.auto_filter.ref = f"A1:{get_column_letter(len(orig_cols))}{1+len(df)}"
ws5.freeze_panes = "A2"

print("[5/5] 원본 데이터 시트 완료")

# ── Save ──
wb.save(OUTPUT_PATH)
print(f"\nExcel 대시보드가 '{OUTPUT_PATH}'에 저장되었습니다.")
